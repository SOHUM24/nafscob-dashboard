import subprocess, sys

# Core
subprocess.run([sys.executable, '-m', 'pip', 'install', 'pandas', 'openpyxl', 'pdfplumber', 'numpy', '--quiet'])

# Optional (better table extraction in many PDFs). If this fails, we will auto-fallback.
# On Windows, Camelot may require extra system deps (Ghostscript). That's OK—fallback still works.
subprocess.run([sys.executable, '-m', 'pip', 'install', 'camelot-py', 'opencv-python', '--quiet'])


import sys
PDF_FOLDER = sys.argv[1]
OUTPUT_FILE = sys.argv[2]
YEAR_MIN, YEAR_MAX = 2000, 2099
PDF_GLOB = '*.pdf'


import re
import numpy as np
import pandas as pd
import pdfplumber

# Optional engine
try:
    import camelot  # type: ignore
    _HAS_CAMELOT = True
except Exception:
    camelot = None
    _HAS_CAMELOT = False


def clean_cell(x):
    if x is None:
        return ''
    s = str(x)
    s = s.replace('\u00a0', ' ')
    s = s.replace('\n', ' ')
    return s.strip()


def normalize_columns(cols):
    out = []
    for c in cols:
        c2 = clean_cell(c)
        c2 = re.sub(r'\s+', ' ', c2)
        out.append(c2)
    return out


def to_number(x):
    """Best-effort numeric conversion (keeps text if not numeric)."""
    s = clean_cell(x)
    if s == '' or s.lower() in {'na', 'n.a', 'n.a.', 'n/d', 'n.d', 'n.d.', '-', '—'}:
        return np.nan
    s2 = s.replace(',', '')
    # handle (123) negatives
    if re.fullmatch(r'\(\s*\d+(?:\.\d+)?\s*\)', s2):
        s2 = '-' + re.sub(r'[()\s]', '', s2)
    # allow percent sign
    if s2.endswith('%'):
        s2 = s2[:-1]
    try:
        return float(s2)
    except Exception:
        return np.nan


def extract_tables_camelot(pdf_path: str):
    """Return list of DataFrames with metadata (best effort)."""
    if not _HAS_CAMELOT:
        return []
    dfs = []
    # try lattice then stream
    for flavor in ['lattice', 'stream']:
        try:
            tables = camelot.read_pdf(pdf_path, pages='all', flavor=flavor)
            for i, t in enumerate(tables):
                df = t.df.copy()
                df.columns = normalize_columns(df.iloc[0].tolist())
                df = df.iloc[1:].reset_index(drop=True)
                df = df.applymap(clean_cell)
                dfs.append({
                    'engine': f'camelot-{flavor}',
                    'page': int(getattr(t, 'page', 0)) if getattr(t, 'page', None) else None,
                    'table_index': i,
                    'df': df,
                })
        except Exception:
            continue
    return dfs


def extract_tables_pdfplumber(pdf_path: str):
    dfs = []
    with pdfplumber.open(pdf_path) as pdf:
        for p_idx, page in enumerate(pdf.pages, start=1):
            try:
                tables = page.extract_tables() or []
            except Exception:
                tables = []
            for t_idx, tbl in enumerate(tables):
                if not tbl or len(tbl) < 2:
                    continue
                df = pd.DataFrame(tbl)
                df = df.applymap(clean_cell)
                # guess header row (row 0)
                df.columns = normalize_columns(df.iloc[0].tolist())
                df = df.iloc[1:].reset_index(drop=True)
                # drop fully empty rows
                df = df.loc[~(df.apply(lambda r: all(clean_cell(v) == '' for v in r), axis=1))].reset_index(drop=True)
                dfs.append({
                    'engine': 'pdfplumber',
                    'page': p_idx,
                    'table_index': t_idx,
                    'df': df,
                })
    return dfs


def extract_all_tables(pdf_path: str):
    """Try Camelot first; fallback to pdfplumber."""
    out = extract_tables_camelot(pdf_path)
    if out:
        return out
    return extract_tables_pdfplumber(pdf_path)


def raw_long_from_df(df: pd.DataFrame, table_id: str):
    rows = []
    for r in range(df.shape[0]):
        for c in range(df.shape[1]):
            rows.append({
                'table_id': table_id,
                'row': r,
                'col': c,
                'column_name': df.columns[c],
                'value': clean_cell(df.iat[r, c]),
            })
    return pd.DataFrame(rows)


def to_tidy(df: pd.DataFrame, year: int, source_file: str, table_id: str):
    """Generic normalization: first column = entity, other columns = metrics."""
    if df.empty or df.shape[1] < 2:
        return pd.DataFrame(columns=['year', 'source_file', 'table_id', 'entity', 'metric', 'value', 'value_num'])

    work = df.copy()
    work.columns = normalize_columns(work.columns)

    entity_col = work.columns[0]
    # if first col header looks like a serial number, use next col
    if re.search(r'^(s\.?no\.?|sl\.?no\.?|sr\.?no\.?|no\.?|#)$', entity_col.strip().lower()):
        if work.shape[1] >= 2:
            entity_col = work.columns[1]

    id_vars = [entity_col]
    value_vars = [c for c in work.columns if c not in id_vars]

    tidy = work.melt(id_vars=id_vars, value_vars=value_vars, var_name='metric', value_name='value')
    tidy.rename(columns={entity_col: 'entity'}, inplace=True)

    tidy['year'] = int(year)
    tidy['source_file'] = os.path.basename(source_file)
    tidy['table_id'] = table_id

    tidy['entity'] = tidy['entity'].map(clean_cell)
    tidy['metric'] = tidy['metric'].map(clean_cell)
    tidy['value'] = tidy['value'].map(clean_cell)
    tidy['value_num'] = tidy['value'].map(to_number)

    # drop empty
    tidy = tidy.loc[~((tidy['entity'] == '') & (tidy['metric'] == '') & (tidy['value'] == ''))].reset_index(drop=True)
    return tidy


print('Generic extraction + normalization helpers defined.')

table_index_rows = []
raw_long_parts = []
tidy_parts = []

for year, pdf_path in PDF_LIST:
    pdf_name = os.path.basename(pdf_path)
    print(f'\n>>> {year} | {pdf_name}')

    extracted = extract_all_tables(pdf_path)
    print(f'  tables_found: {len(extracted)} (engine preference: camelot→pdfplumber)')

    for t in extracted:
        df = t['df']
        engine = t.get('engine')
        page = t.get('page')
        t_idx = int(t.get('table_index', 0))

        table_id = f"{year}__{Path(pdf_name).stem}__p{page if page else 'NA'}__t{t_idx}"

        table_index_rows.append({
            'table_id': table_id,
            'year': year,
            'source_file': pdf_name,
            'page': page,
            'table_index_on_page': t_idx,
            'engine': engine,
            'rows': int(df.shape[0]),
            'cols': int(df.shape[1]),
            'columns': ' | '.join([clean_cell(c) for c in df.columns]),
        })

        raw_long_parts.append(raw_long_from_df(df, table_id))
        tidy_parts.append(to_tidy(df, year=year, source_file=pdf_path, table_id=table_id))

TABLE_INDEX = pd.DataFrame(table_index_rows)
RAW_TABLES = pd.concat(raw_long_parts, ignore_index=True) if raw_long_parts else pd.DataFrame()
TIDY = pd.concat(tidy_parts, ignore_index=True) if tidy_parts else pd.DataFrame()

print('\n✓ Extraction complete')
print('TABLE_INDEX:', TABLE_INDEX.shape)
print('RAW_TABLES:', RAW_TABLES.shape)
print('TIDY:', TIDY.shape)

from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

CLR_HEADER = 'FF1F3864'
CLR_SUBHEAD = 'FF2E75B6'
CLR_WHITE = 'FFFFFFFF'


def write_df(ws, df: pd.DataFrame, title: Optional[str] = None):
    ws.sheet_view.showGridLines = False

    start_row = 1
    if title:
        ws.append([title])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, len(df.columns)))
        c = ws.cell(1, 1)
        c.font = Font(bold=True, size=12, color=CLR_WHITE)
        c.fill = PatternFill('solid', start_color=CLR_HEADER)
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 22
        start_row = 2

    if df is None or df.empty:
        ws.append(['No data'])
        return

    # header
    ws.append(list(df.columns))
    for j in range(1, len(df.columns) + 1):
        cell = ws.cell(start_row, j)
        cell.font = Font(bold=True, color=CLR_WHITE)
        cell.fill = PatternFill('solid', start_color=CLR_SUBHEAD)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[start_row].height = 26

    # data
    for row in dataframe_to_rows(df, index=False, header=False):
        ws.append(row)

    # widths
    for j in range(1, len(df.columns) + 1):
        col = get_column_letter(j)
        max_len = 10
        for i in range(1, min(ws.max_row, 5000) + 1):
            v = ws.cell(i, j).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[col].width = min(45, max_len + 2)

    ws.freeze_panes = ws['A' + str(start_row + 1)]


print('Excel writer ready.')

def build_yoy(tidy: pd.DataFrame):
    if tidy.empty:
        return pd.DataFrame(columns=['entity','metric','year','value_num','prev_year','prev_value','delta','yoy_pct'])

    df = tidy.copy()
    df = df.loc[~df['value_num'].isna()].copy()
    df = df.sort_values(['entity', 'metric', 'year'])

    df['prev_year'] = df.groupby(['entity','metric'])['year'].shift(1)
    df['prev_value'] = df.groupby(['entity','metric'])['value_num'].shift(1)
    df['delta'] = df['value_num'] - df['prev_value']
    df['yoy_pct'] = np.where(df['prev_value'].isna() | (df['prev_value'] == 0), np.nan, (df['delta'] / df['prev_value']) * 100.0)

    return df[['entity','metric','year','value_num','prev_year','prev_value','delta','yoy_pct','source_file','table_id']]


def build_pivots(tidy: pd.DataFrame):
    if tidy.empty:
        return {}, {}
    base = tidy.loc[~tidy['value_num'].isna()].copy()

    # metric × year (All entities summed)
    p_metric_year = base.pivot_table(
        index='metric',
        columns='year',
        values='value_num',
        aggfunc='sum'
    ).reset_index()

    # entity × year (All metrics summed)
    p_entity_year = base.pivot_table(
        index='entity',
        columns='year',
        values='value_num',
        aggfunc='sum'
    ).reset_index()

    return p_metric_year, p_entity_year


YOY = build_yoy(TIDY)
PIVOT_METRIC_YEAR, PIVOT_ENTITY_YEAR = build_pivots(TIDY)

print('YOY:', YOY.shape)
print('PIVOT_METRIC_YEAR:', PIVOT_METRIC_YEAR.shape)
print('PIVOT_ENTITY_YEAR:', PIVOT_ENTITY_YEAR.shape)

# ── Write the workbook ────────────────────────────────────────────────────────

out_path = str(Path(PDF_FOLDER) / OUTPUT_FILE)

wb = Workbook()
ws0 = wb.active
ws0.title = 'TABLE_INDEX'
write_df(ws0, TABLE_INDEX, title='Extracted table index')

ws = wb.create_sheet('RAW_TABLES')
write_df(ws, RAW_TABLES, title='Raw extracted tables (long form)')

ws = wb.create_sheet('TIDY')
write_df(ws, TIDY, title='Tidy dataset for visualization')

ws = wb.create_sheet('YOY')
write_df(ws, YOY, title='Year-over-year analysis')

ws = wb.create_sheet('PIVOT_METRIC_YEAR')
write_df(ws, PIVOT_METRIC_YEAR, title='Pivot: metric × year (sum)')

ws = wb.create_sheet('PIVOT_ENTITY_YEAR')
write_df(ws, PIVOT_ENTITY_YEAR, title='Pivot: entity × year (sum)')

wb.save(out_path)
print(f'✅ Saved Excel → {out_path}')

from openpyxl import load_workbook

wb_check = load_workbook(out_path, read_only=True)
print(f'Workbook: {out_path}')
print(f'Total sheets: {len(wb_check.sheetnames)}')
print('Sheets:', wb_check.sheetnames)

for name in wb_check.sheetnames:
    ws = wb_check[name]
    print(f'{name:<20} rows={ws.max_row:<8} cols={ws.max_column:<5}')

wb_check.close()
print('✅ Verification complete.')

# Quick preview: top metrics and entities

display(TABLE_INDEX.head(10))

display(TIDY.head(20))

# Most common metrics captured
if not TIDY.empty:
    display(TIDY['metric'].value_counts().head(25))

# Quick preview: YoY biggest changes (absolute)
if not YOY.empty:
    top_moves = YOY.dropna(subset=['delta']).sort_values('delta', ascending=False).head(20)
    display(top_moves)
else:
    print('YOY is empty — likely no numeric columns were detected. Check RAW_TABLES/TIDY and adjust normalization if needed.')

